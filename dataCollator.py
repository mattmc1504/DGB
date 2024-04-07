# Program that takes rough excel data given from proprietary software and collates it under relevant groups
# printed and filtered, ready to sort and compare

import openpyxl 
wb = openpyxl.load_workbook("Product Sales Feb.xlsx") 
wb2 = openpyxl.Workbook()

for sheet in wb.worksheets:
    row = sheet.max_row 
    column = sheet.max_column 
    sheet2 = wb2.create_sheet (title = " ")
    for i in range(5, row + 1): 
        item = sheet.cell(row=i, column=5)
        net_sales = sheet.cell(row=i, column=9).value
        percent = sheet.cell(row=i, column=10).value     
# Looks forwards and backwards at product names to tell when to move on to the next product or collate multiple lines that share the same product name
        item_following = sheet.cell(row=i+1, column=5).value
        item_preceding = sheet.cell(row=i-1, column=5).value
        if item.value == None:
            continue       
        if item.value == "":
            continue   
        if item_preceding == item.value:
            continue
# Skips category total lines and specifies just the individual products       
        if item.value == "SubTotal":
            continue
        if item_following == item.value:
            for x in range(1, row-i+1):
                item_iteration = sheet.cell(row=i+x, column=5)
                if item_iteration.value != item.value:
                    break
                if sheet.cell(row=i+x, column=9).value == None:
                    percent += sheet.cell(row=i+x, column=10).value
                    continue    
                elif sheet.cell(row=i+x, column=10).value == None:     
                    net_sales += sheet.cell(row=i+x, column=9).value
                    continue                    
                else:
                    net_sales += sheet.cell(row=i+x, column=9).value                
                    percent += sheet.cell(row=i+x, column=10).value     
# Prints data into newly created workbook, with a new sheet for each in the base workbook                    
        c1 = sheet2.cell(row = i, column = 6)
        c1.value = str(item.value)
        c2 = sheet2.cell(row = i, column = 7)
        c2.value = (net_sales)
        c3 = sheet2.cell(row = i, column = 8)
        if isinstance(percent, str) == True:
            continue
        else:
            c3.value = (percent)
        
wb.save ("testExcel.xlsx")
wb2.save ("CollatedFeb.xlsx")