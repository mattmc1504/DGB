# Program that takes collated data and produces percentage change figures
# Based on another set of data

import openpyxl 
wb2 = openpyxl.load_workbook("CollatedMarch.xlsx")

# Compares the item names in the 2 sets of data and when a match is found, produces % change figures for each of the 2 columns of data associated with it
for sheet in wb2.worksheets:
    row = sheet.max_row 
    column = sheet.max_column 
    for i in range(1, row + 1): 
        item1 = sheet.cell(row=i, column=6).value
        net_sales1 = sheet.cell(row=i, column=7).value
        percent1 = sheet.cell(row=i, column=8).value         
        for x in range(1, row + 1):
            item2 = sheet.cell(row=x, column=2).value
            net_sales2 = sheet.cell(row=x, column=3).value
            percent2 = sheet.cell(row=x, column=4).value   
            if item2 == None:
                continue                               
            if item1 == item2:
                if net_sales2 != 0 and isinstance(net_sales2, str) == False:
                    sales_change = (net_sales1 / net_sales2)
                    c1 = sheet.cell(row = i, column = 9)
                    c1.value = (float(sales_change)-1)
                else:
                    c1 = sheet.cell(row = i, column = 9)
                    c1.value = (net_sales1)                
                if net_sales2 != 0 and isinstance(net_sales2, str) == False:                
                    percent_change = (percent1 - percent2)
                    c2 = sheet.cell(row = i, column = 10)
                    c2.value = (float(percent_change)/100)
            
# For new products, fills in the empty spaces due to lack of comparison with the "NEW" descriptor            
for sheet in wb2.worksheets:
    row = sheet.max_row 
    column = sheet.max_column 
    for i in range(1, row + 1):
        c1 = sheet.cell(row = i, column = 9)
        c2 = sheet.cell(row = i, column = 10)        
        if c1.value == None:
            c1.value = ("NEW")    
        if c2.value == None:
            c2.value = ("NEW") 
        else:
            continue    

# Applies filter to the area to allow sorting of the compared product list                
    filters = sheet.auto_filter
    filters.ref = "F2:J2000"

wb2.save ("CollatedMarch.xlsx")