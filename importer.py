# Program designed to take a list of given product price change data, sort it into a readable format and allow transposition into our software

import openpyxl 
wb = openpyxl.load_workbook("MC Price Changes.xlsx") 
sheet = wb["Sheet2"] 
sheetSplit = wb.create_sheet (title = "split2")

row = sheet.max_row 
column = sheet.max_column 

# Data is copied into excel in one column with each cell being a long string of information
# This is broken into a list, with the relevant parts taken and printed into seperate columns in a new sheet
for i in range(1, row + 1): 
    item = str((sheet.cell(row=i, column=2)).value)
    split = item.split (" ")
    if len(split) > 6:
        c1 = sheetSplit.cell(row = i, column = 2)
        c1String = str(split [0])
        c1StringFormat = c1String [3:]
        c1.value = c1String
        c2 = sheetSplit.cell(row = i, column = 3)
        c2String = str(split [-1])
        c2StringFormat = c2String [1:]
        c2.value = c2StringFormat

wb.save ("MC Price Changes.xlsx")

# Takes this seperated out data from above and compares to the import file
wb1 = openpyxl.load_workbook("priceImport.xlsx") 
wb2 = openpyxl.load_workbook("MC Price Changes.xlsx") 
sheet1 = wb1["Sheet1"]
sheet2 = wb2["Current Prices"]
row1 = sheet1.max_row
row2 = sheet2.max_row
column1 = sheet1.max_column
column2 = sheet2.max_column

# First matches the data to the relevant supplier
# Then compares the column from each sheet detailing the product code
# If there is a match the new price is copied into the import file in the relevant cell
for i2 in range (1, row2 + 1):
    if sheet2.cell(row = i2, column = 3).value == None:
        continue
    else:
        for i1 in range (1, row1 + 1):
            if sheet1.cell(row = i1, column = 1).value == "Matthew Clark":
                if sheet1.cell(row = i1, column = 8).value == None:
                    continue
                if str(sheet1.cell(row = i1, column = 8).value) == str(sheet2.cell(row = i2, column = 3).value):
                    newCost = sheet1.cell(row = i1, column = 10)
                    newCost.value = float(sheet2.cell(row = i2, column = 4).value)                                             
            else:
                continue

# Specific bit of code for suppliers whose product codes start with 0's, as this is removed by excel in number formatting
# Finds the codes where this has occurred and adds the 0's back to the start, reprinting as a string so they are not lost again
for i1 in range (1, row1 + 1):
    if sheet1.cell(row = i1, column = 1).value == "Matthew Clark":
        newCode = sheet1.cell(row = i1, column = 8)
        if sheet1.cell(row = i1, column = 8) == None:
            continue
        if len(str(sheet1.cell(row = i1, column = 8).value)) == 4:
            newCode.value = "0000" + str(sheet1.cell(row = i1, column = 8).value)        
        if len(str(sheet1.cell(row = i1, column = 8).value)) == 5:
            newCode.value = "000" + str(sheet1.cell(row = i1, column = 8).value)
        if len(str(sheet1.cell(row = i1, column = 8).value)) == 6:
            newCode.value = "00" + str(sheet1.cell(row = i1, column = 8).value)  
        if len(str(sheet1.cell(row = i1, column = 8).value)) == 7:
            newCode.value = "0" + str(sheet1.cell(row = i1, column = 8).value)  
        else:
            continue  

print ("Sheet comparison complete.")

wb2.save ("MC Price Changes.xlsx")
wb1.save ("priceImport.xlsx")